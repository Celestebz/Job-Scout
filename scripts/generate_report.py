#!/usr/bin/env python3
"""
Job Scout Report Generator
生成XLSX格式岗位报告

用途：根据搜索结果生成Excel格式的岗位报告
期望薪资：12-18K
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json
import os

# ===== 配置区域 =====
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
RESULT_DIR = os.path.join(PROJECT_ROOT, 'result')
JOB_HISTORY_FILE = os.path.join(PROJECT_ROOT, 'job_history.json')

# 新发现的岗位数据（12-18K期望薪资搜索结果）
NEW_JOBS = [
    {
        "rank": 1,
        "company": "路特创新（A轮）",
        "title": "Reddit/Quora海外社区运营专员",
        "salary": "12-26K",
        "location": "深圳-坂田",
        "match_score": 90,
        "jd": "岗位职责：\n1. 负责Reddit/Quora等海外社区运营\n2. 社区内容策划和用户互动\n3. 社区规则维护和危机处理\n4. 数据分析和优化\n\n任职要求：\n1. 本科及以上学历，3年以上经验\n2. 英语流利，熟悉海外社区文化\n3. 有社区运营经验\n4. 良好的沟通能力",
        "apply_url": "猎聘",
        "notes": "NEW - 12-26K薪资范围，A轮公司，3年经验要求，英语专八优势明显",
        "search_date": "2026-03-03"
    },
    {
        "rank": 2,
        "company": "绿联科技（上市公司）",
        "title": "海外市场经理",
        "salary": "12-18K·14薪",
        "location": "深圳",
        "match_score": 92,
        "jd": "岗位职责：\n1. 负责海外市场拓展和运营\n2. 海外渠道建设和维护\n3. 市场活动策划和执行\n4. 销售数据分析\n\n任职要求：\n1. 本科及以上学历，3-5年经验\n2. 有海外销售或市场经验\n3. 英语能力良好\n4. 沟通协调能力强",
        "apply_url": "猎聘",
        "notes": "NEW - 上市公司，12-18K·14薪，3年海外销售经验完美匹配",
        "search_date": "2026-03-03"
    },
    {
        "rank": 3,
        "company": "深圳优维尔科技",
        "title": "海外社媒运营专员（TikTok）",
        "salary": "12-18K·14薪",
        "location": "深圳-西乡",
        "match_score": 88,
        "jd": "岗位职责：\n1. 负责TikTok等海外社媒平台运营\n2. 内容策划和发布\n3. 粉丝增长和互动\n4. 数据分析和优化\n\n任职要求：\n1. 本科及以上学历，1-3年经验\n2. 熟悉TikTok等海外社媒平台\n3. 英语能力良好\n4. 有创意和执行力",
        "apply_url": "猎聘",
        "notes": "NEW - 12-18K·14薪，1-3年经验，TikTok运营方向",
        "search_date": "2026-03-03"
    },
    {
        "rank": 4,
        "company": "有生之颜日化",
        "title": "独立站运营",
        "salary": "12-18K·13薪",
        "location": "深圳-福田区",
        "match_score": 85,
        "jd": "岗位职责：\n1. 负责独立站日常运营\n2. 产品上架和优化\n3. 数据分析和销售提升\n4. 客服和售后支持\n\n任职要求：\n1. 有独立站运营经验\n2. 了解跨境电商\n3. 数据分析能力\n4. 学习能力强",
        "apply_url": "猎聘",
        "notes": "NEW - 12-18K·13薪，独立站运营，福田区位置",
        "search_date": "2026-03-03"
    },
    {
        "rank": 5,
        "company": "宝安区某公司",
        "title": "海外红人营销专员",
        "salary": "15-30K",
        "location": "深圳-宝安区",
        "match_score": 86,
        "jd": "岗位职责：\n1. 负责海外红人/KOL资源拓展\n2. 红人合作洽谈和管理\n3. 营销活动策划和执行\n4. 效果数据分析和优化\n\n任职要求：\n1. 有KOL合作经验\n2. 英语流利\n3. 沟通能力强\n4. 有创意思维",
        "apply_url": "猎聘",
        "notes": "NEW - 15-30K薪资，KOL运营经验完美匹配",
        "search_date": "2026-03-03"
    },
    {
        "rank": 6,
        "company": "深圳科思明德医疗科技",
        "title": "小语种海外销售",
        "salary": "12-18K·14薪",
        "location": "深圳-龙华区",
        "match_score": 80,
        "jd": "岗位职责：\n1. 负责海外市场销售工作\n2. 客户开发和维护\n3. 销售目标达成\n4. 市场信息收集\n\n任职要求：\n1. 本科及以上学历，1-3年经验\n2. 英语及小语种能力\n3. 有销售经验\n4. 沟通能力强",
        "apply_url": "猎聘",
        "notes": "NEW - 12-18K·14薪，医疗器械行业，双休",
        "search_date": "2026-03-03"
    },
    {
        "rank": 7,
        "company": "慕迪贸易（C轮）",
        "title": "品牌与市场运营经理（欧洲电商出海）",
        "salary": "10-15K",
        "location": "深圳-南山区",
        "match_score": 85,
        "jd": "岗位职责：\n1. 负责欧洲电商出海运营\n2. 品牌建设和市场推广\n3. 渠道拓展和管理\n4. 数据分析\n\n任职要求：\n1. 本科及以上学历，1-3年经验\n2. 有电商或出海经验\n3. 英语能力良好\n4. 市场营销思维",
        "apply_url": "猎聘",
        "notes": "NEW - C轮公司，10-15K，欧洲电商出海方向，可线上办公",
        "search_date": "2026-03-03"
    },
    {
        "rank": 8,
        "company": "上海珺至贸易",
        "title": "欧洲电商出海渠道拓展负责人",
        "salary": "12-20K",
        "location": "深圳-龙岗区",
        "match_score": 83,
        "jd": "岗位职责：\n1. 负责欧洲市场渠道拓展\n2. 渠道关系维护\n3. 销售目标达成\n4. 市场分析\n\n任职要求：\n1. 本科及以上学历，3年以上经验\n2. 有渠道拓展经验\n3. 英语能力良好\n4. 可线上办公",
        "apply_url": "猎聘",
        "notes": "NEW - 12-20K，欧洲市场，可线上办公",
        "search_date": "2026-03-03"
    },
    {
        "rank": 9,
        "company": "深圳泽立方信息科技",
        "title": "TikTok运营管培生（2026届）",
        "salary": "面议",
        "location": "深圳",
        "match_score": 82,
        "jd": "岗位职责：\n1. TikTok账号运营和内容策划\n2. 短视频制作\n3. 数据分析\n4. 管培生轮岗学习\n\n任职要求：\n1. 2026届应届毕业生\n2. 对短视频运营感兴趣\n3. 学习能力强\n4. 团队协作能力",
        "apply_url": "校招官网",
        "notes": "NEW - 2026届校招，快时尚品牌出海，管培生项目",
        "search_date": "2026-03-03"
    },
    {
        "rank": 10,
        "company": "安克创新",
        "title": "Marketing Specialist（英语专八要求）",
        "salary": "6-11K",
        "location": "深圳-新安",
        "match_score": 88,
        "jd": "岗位职责：\n1. 负责海外市场营销工作\n2. 品牌推广和活动策划\n3. 市场数据分析\n4. 跨部门协作\n\n任职要求：\n1. 本科及以上学历，4年以上经验\n2. **英语专八必须**\n3. 有市场营销经验\n4. 出海的消费电子公司",
        "apply_url": "猎聘",
        "notes": "NEW - 知名出海公司Anker，英语专八硬性要求，6-11K（但品牌好）",
        "search_date": "2026-03-03"
    },
    {
        "rank": 11,
        "company": "某B轮智能硬件公司",
        "title": "社媒与KOL营销专员",
        "salary": "15-25K·15薪",
        "location": "深圳-南山区",
        "match_score": 87,
        "jd": "岗位职责：\n1. 负责海外社媒和KOL营销\n2. 内容策划和投放\n3. 红人合作管理\n4. 效果数据优化\n\n任职要求：\n1. 本科及以上学历\n2. 有社媒或KOL经验\n3. 英语能力良好\n4. 数据驱动思维",
        "apply_url": "猎聘",
        "notes": "NEW - B轮公司，15-25K·15薪，南山区位置",
        "search_date": "2026-03-03"
    },
    {
        "rank": 12,
        "company": "某深圳在线社交/媒体公司",
        "title": "海外KOL运营",
        "salary": "25-45K·13薪",
        "location": "深圳",
        "match_score": 75,
        "jd": "岗位职责：\n1. 负责海外KOL运营策略\n2. KOL资源拓展和管理\n3. 营销活动策划\n4. 数据分析\n\n任职要求：\n1. 有KOL运营经验\n2. 英语流利\n3. 有管理能力\n4. 行业了解深入",
        "apply_url": "猎聘",
        "notes": "NEW - 25-45K高薪，但可能要求更高级经验，可冲刺尝试",
        "search_date": "2026-03-03"
    },
]


def create_xlsx_report(jobs, output_path):
    """创建XLSX格式的岗位报告"""

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "岗位侦察报告"

    # 定义样式
    header_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font = Font(name='Arial', size=14)
    data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    jd_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = ['排名', '公司', '岗位', '薪资', '地点', '匹配度', '岗位描述(JD)', '投递方式', '备注', '搜索日期']
    ws.append(headers)

    # 设置表头样式
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # 写入数据
    row_num = 2
    for job in jobs:
        ws.append([
            job['rank'],
            job['company'],
            job['title'],
            job['salary'],
            job['location'],
            job['match_score'],
            job['jd'],
            job['apply_url'],
            job['notes'],
            job['search_date']
        ])

        # 设置数据行样式
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = border
            if col_num == 7:  # JD列左对齐
                cell.alignment = jd_alignment
            else:
                cell.alignment = data_alignment

        row_num += 1

    # 设置列宽
    column_widths = {
        'A': 6,   # 排名
        'B': 25,  # 公司
        'C': 30,  # 岗位
        'D': 18,  # 薪资
        'E': 18,  # 地点
        'F': 10,  # 匹配度
        'G': 60,  # JD
        'H': 25,  # 投递方式
        'I': 50,  # 备注
        'J': 15,  # 搜索日期
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置行高
    for row in range(1, row_num):
        ws.row_dimensions[row].height = 35

    # 保存文件
    wb.save(output_path)
    return output_path


def main():
    """主函数"""

    # 确保result目录存在
    os.makedirs(RESULT_DIR, exist_ok=True)

    # 生成报告文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'job_scout_王宝珍_12-18K_{timestamp}.xlsx'
    output_path = os.path.join(RESULT_DIR, filename)

    # 生成报告
    create_xlsx_report(NEW_JOBS, output_path)

    print(f"✅ 报告已生成: {output_path}")
    print(f"📊 共 {len(NEW_JOBS)} 个新岗位")
    print(f"💰 期望薪资: 12-18K")
    print(f"📅 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == '__main__':
    main()
